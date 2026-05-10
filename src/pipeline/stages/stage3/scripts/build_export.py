from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pipeline.schemas.stage3 import PartRecord

_HEADERS = [
    "Assembly ID",
    "Part Name",
    "OEM Part Number",
    "Brand",
    "Description",
    "Quantity",
    "Unit",
    "Source URL",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_UNVERIFIED_FILL = PatternFill("solid", fgColor="FFF2CC")


def build_excel_export(parts: list[PartRecord], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parts"

    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, part in enumerate(parts, start=2):
        row_values = [
            part.assembly_id,
            part.part_name,
            part.oem_part_number,
            part.brand,
            part.description,
            part.quantity,
            part.unit,
            part.source_url,
        ]
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if value == "UNVERIFIED":
                cell.fill = _UNVERIFIED_FILL

    for col_idx in range(1, len(_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    wb.save(output_path)
